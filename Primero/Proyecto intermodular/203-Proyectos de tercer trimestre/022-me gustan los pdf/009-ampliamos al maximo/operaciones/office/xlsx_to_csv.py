#!/usr/bin/env python3

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is not installed. Install it with: sudo apt install python3-openpyxl")
    sys.exit(1)


def ensure_dir(path: Path):
    path.mkdir(parents=True, exist_ok=True)


def collect_xlsx(inputs):
    files = []
    for item in inputs:
        p = Path(item)
        if p.is_file() and p.suffix.lower() == ".xlsx":
            files.append(p.resolve())
        elif p.is_dir():
            for child in sorted(p.iterdir()):
                if child.is_file() and child.suffix.lower() == ".xlsx":
                    files.append(child.resolve())
    return files


def safe_cell(value):
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def xlsx_to_csv(input_path: Path, output_dir: Path):
    wb = load_workbook(input_path, data_only=True)

    output_files = []

    for sheet in wb.worksheets:
        csv_name = f"{input_path.stem}_{sheet.title}.csv"
        csv_name = csv_name.replace(" ", "_")

        output_path = output_dir / csv_name

        with open(output_path, "w", encoding="utf-8") as f:
            for row in sheet.iter_rows(values_only=True):
                line = ",".join(safe_cell(cell) for cell in row)
                f.write(line + "\n")

        output_files.append(output_path)

    return output_files


def main():
    """
    Usage:
        python3 xlsx_to_csv.py <output_dir> <file_or_folder> [more...]

    Requires:
        sudo apt install python3-openpyxl
    """

    if len(sys.argv) < 3:
        print("Usage:")
        print("  python3 xlsx_to_csv.py <output_dir> <file_or_folder> [more_files_or_folders...]")
        sys.exit(1)

    output_dir = Path(sys.argv[1]).resolve()
    inputs = sys.argv[2:]

    ensure_dir(output_dir)

    files = collect_xlsx(inputs)

    if not files:
        print("ERROR: No XLSX files found.")
        sys.exit(1)

    processed = 0
    errors = []

    for f in files:
        try:
            outputs = xlsx_to_csv(f, output_dir)
            processed += len(outputs)
        except Exception as e:
            errors.append(f"{f}: {e}")

    report = output_dir / "report.txt"
    with open(report, "w", encoding="utf-8") as r:
        r.write("XLSX TO CSV REPORT\n")
        r.write("==================\n\n")
        r.write(f"Generated CSV files: {processed}\n")
        r.write(f"Errors: {len(errors)}\n\n")

        if errors:
            r.write("ERRORS\n------\n")
            for e in errors:
                r.write(e + "\n")

    if processed == 0:
        print("ERROR: No files converted.")
        print(f"REPORT: {report}")
        sys.exit(1)

    print("OK")
    print(f"Generated CSV files: {processed}")
    print(f"Errors: {len(errors)}")
    print(f"OUTPUT_DIR: {output_dir}")
    print(f"REPORT: {report}")


if __name__ == "__main__":
    main()
